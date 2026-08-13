"""Generadores de archivos oficiales de CLARA+.

Los generadores reciben declaraciones ya cargadas desde la base de datos para
mantener la lógica de presentación separada de las rutas HTTP.
"""

from io import BytesIO
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib import colors
from reportlab.lib.pagesizes import portrait
from reportlab.lib.units import cm
from reportlab.pdfgen import canvas
from reportlab.graphics.barcode import qr
from reportlab.graphics.shapes import Drawing


def _style_header(ws) -> None:
    fill = PatternFill("solid", fgColor="1F3B5B")
    for cell in ws[1]:
        cell.fill = fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def _record_row(record) -> list:
    laboratorio = record.laboratorio
    return [
        # El formato oficial imprime el "Código" legible; el identificador
        # técnico se conserva aparte porque aquel no es único entre registros.
        record.codigo_formato or record.codigo_residuo,
        record.fecha.isoformat(),
        laboratorio.dependencia.nombre if laboratorio and laboratorio.dependencia else "",
        laboratorio.nombre if laboratorio else "",
        record.actividad,
        record.origen,
        record.responsable,
        record.descripcion,
        ", ".join(i.nombre_quimico for i in record.insumos),
        record.estado_fisico,
        # Cantidad y unidad tal como las pide el formato oficial; el pesaje va
        # aparte porque es la evidencia, no la cifra declarada.
        record.cantidad,
        record.unidad,
        record.tara_g,
        record.peso_bruto_g,
        record.peso_neto_g,
        record.ph,
        record.categoria.nombre if record.categoria else "",
        record.categoria.clase_sunat if record.categoria else "",
        record.categoria.clase_basilea if record.categoria else "",
        record.confianza,
        "Sí" if record.escalar_csbqr else "No",
    ]


def workbook_declaracion(records: Iterable, month: int, year: int) -> BytesIO:
    record_list = list(records)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Declaración"
    headers = [
        "ID residuo", "Fecha", "Dependencia", "Laboratorio", "Actividad", "Origen",
        "Responsable", "Descripción", "Insumos", "Estado físico", "Cantidad",
        "Unidad", "Tara (g)", "Peso bruto (g)", "Peso neto (g)", "pH", "Categoría",
        "Clase declaración", "Basilea", "Confianza", "Escalar CSBQR",
    ]
    sheet.append(headers)
    for record in record_list:
        sheet.append(_record_row(record))
    _style_header(sheet)
    sheet.title = f"Declaración {month:02d}-{year}"
    for column in sheet.columns:
        letter = column[0].column_letter
        sheet.column_dimensions[letter].width = min(max(len(str(column[0].value or "")) + 2, 12), 36)

    totals = workbook.create_sheet("Totales")
    totals.append(["Período", f"{month:02d}/{year}"])
    totals.append(["Cantidad de residuos", len(record_list)])
    # Solo se totaliza lo declarado en kilogramos: sumar kg con litros no
    # produce ninguna magnitud.
    kilos = sum(r.cantidad for r in record_list if r.unidad == "Kg")
    litros = sum(r.cantidad for r in record_list if r.unidad == "L")
    totals.append(["Cantidad total declarada en kg", round(kilos, 4)])
    totals.append(["Cantidad total declarada en L", round(litros, 4)])
    _style_header(totals)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def workbook_traslado(records: Iterable, transportista: str) -> BytesIO:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Traslado"
    sheet.append(["ID residuo", "Fecha", "Laboratorio origen", "Responsable", "Categoría", "Cantidad", "Unidad", "Transportista"])
    record_list = list(records)
    for record in record_list:
        sheet.append([
            record.codigo_residuo,
            record.fecha.isoformat(),
            record.laboratorio.nombre if record.laboratorio else "",
            record.responsable,
            record.categoria.nombre if record.categoria else "",
            record.cantidad,
            record.unidad,
            transportista,
        ])
    _style_header(sheet)
    for column in sheet.columns:
        sheet.column_dimensions[column[0].column_letter].width = 24

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def etiqueta_pdf(record) -> BytesIO:
    width, height = 10 * cm, 15 * cm
    output = BytesIO()
    pdf = canvas.Canvas(output, pagesize=portrait((width, height)))

    # Franja institucional y título.
    pdf.setFillColor(colors.HexColor("#EE7623"))
    pdf.rect(0, height - 1.35 * cm, width, 1.35 * cm, fill=1, stroke=0)
    pdf.setFillColor(colors.white)
    pdf.setFont("Helvetica-Bold", 15)
    pdf.drawCentredString(width / 2, height - 0.82 * cm, "RESIDUO PELIGROSO")

    pdf.setFillColor(colors.HexColor("#1F3B5B"))
    pdf.setFont("Helvetica-Bold", 10)
    y = height - 2.0 * cm
    pdf.drawString(0.65 * cm, y, record.categoria.nombre if record.categoria else "Residuo peligroso")
    y -= 0.55 * cm
    pdf.setFont("Helvetica", 8.5)
    pdf.drawString(0.65 * cm, y, f"Código: {record.codigo_formato or record.codigo_residuo}")
    if record.registro:
        y -= 0.42 * cm
        pdf.drawString(0.65 * cm, y, f"Registro: {record.registro.codigo}")
    y -= 0.42 * cm
    pdf.drawString(0.65 * cm, y, f"Grupo: {record.categoria.grupo_compatibilidad if record.categoria else 'AISLAR'}")
    y -= 0.42 * cm
    pdf.drawString(0.65 * cm, y, f"Cantidad: {record.cantidad:g} {record.unidad}")
    y -= 0.42 * cm
    pdf.drawString(0.65 * cm, y, f"Fecha: {record.fecha}")

    # El bloque de texto avanza según las líneas que realmente ocupa. Con un
    # desplazamiento fijo, una descripción larga se imprimía encima del bloque
    # siguiente y tapaba los pictogramas.
    ALTO_LINEA = 0.35 * cm
    LIMITE_INFERIOR = 3.15 * cm  # Espacio reservado para el QR y el pie.

    def bloque(titulo: str, contenido: str, y_actual: float, ancho: int = 58) -> float:
        lineas = [contenido[i:i + ancho] for i in range(0, len(contenido), ancho)] or [""]
        y_actual -= 0.6 * cm
        pdf.setFont("Helvetica-Bold", 8.5)
        pdf.drawString(0.65 * cm, y_actual, titulo)
        pdf.setFont("Helvetica", 8)
        for linea in lineas:
            y_actual -= ALTO_LINEA
            if y_actual < LIMITE_INFERIOR:
                # Antes que invadir el QR, se corta y se avisa en la etiqueta.
                pdf.drawString(0.65 * cm, y_actual + ALTO_LINEA / 2, "[...] ver ficha completa")
                break
            pdf.drawString(0.65 * cm, y_actual, linea)
        return y_actual

    y = bloque("Descripción", record.descripcion or "Sin descripción", y)

    pictogramas = ", ".join(p.codigo_pictograma for p in record.pictogramas) or "Según evaluación"
    y = bloque("Pictogramas GHS", pictogramas, y)

    # Incompatibilidades reales de la categoría. Antes decía "Consultar matriz de
    # compatibilidad", que obliga a ir a buscar el dato justo cuando hace falta
    # tenerlo a la vista, con el envase en la mano.
    incompatibles = (record.categoria.no_mezclar_con if record.categoria else "") or ""
    y = bloque(
        "NO MEZCLAR CON",
        incompatibles.strip() or "Consultar al CSBQR antes de agrupar",
        y,
    )

    # QR con el identificador del registro.
    qr_code = qr.QrCodeWidget(record.codigo_residuo)
    bounds = qr_code.getBounds()
    size = 2.25 * cm
    drawing = Drawing(size, size, transform=[size / (bounds[2] - bounds[0]), 0, 0, size / (bounds[3] - bounds[1]), 0, 0])
    drawing.add(qr_code)
    drawing.drawOn(pdf, width - size - 0.65 * cm, 0.65 * cm)
    pdf.setFont("Helvetica", 6.5)
    pdf.drawString(0.65 * cm, 0.78 * cm, "Mantener cerrado y en contención secundaria")

    pdf.showPage()
    pdf.save()
    output.seek(0)
    return output
